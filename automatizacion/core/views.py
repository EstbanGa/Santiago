from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
import openpyxl
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db.models import Q, Sum
from django.core.exceptions import ValidationError
from django.core.serializers.json import DjangoJSONEncoder
from django.utils.safestring import mark_safe
from django.contrib import messages
from django.utils.dateparse import parse_date

from .models import Registro, Cliente, Proveedor, Maquina
from .forms import RegistroForm, MaquinaForm

from decimal import Decimal
from datetime import datetime, date, timedelta
import json



def index(request):
    return render(request, 'index.html')

# Vista para la sección de Tesorería
def vista_tesoreria(request):
    # Aquí puedes agregar la lógica para tu página de tesorería
    return render(request, 'tesoreria.html')

def vista_maquinaria(request):
    query = request.GET.get('q', '')
    
    if query:
        maquinas = Maquina.objects.filter(
            Q(nombre__icontains=query) | 
            Q(numero_serie__icontains=query)
        )
    else:
        maquinas = Maquina.objects.all()
    
    # Cambia esta línea para usar el template correcto
    return render(request, 'maquinaria.html', {  # o el nombre que tengas
        'maquinas': maquinas,
        'query': query
    })

def ver_detalle_maquina(request, id):
    """Vista para ver los detalles de una máquina específica"""
    maquina = get_object_or_404(Maquina, id=id)
    return render(request, 'detalle.html', {
        'maquina': maquina
    })

def vista_crear_maquinaria(request):
    if request.method == 'POST':
        form = MaquinaForm(request.POST)
        if form.is_valid():
            form.save() # 👈 ¡Esto guarda todos los datos automáticamente!
            return redirect('maquinaria')
    else:
        form = MaquinaForm() # Crea un formulario vacío

    context = {
        'form': form
    }
    return render(request, 'crear_maquinaria.html', context)

def editar_maquina(request, id):
    """
    Vista para editar una máquina existente
    """
    maquina = get_object_or_404(Maquina, id=id)
    
    if request.method == 'POST':
        form = MaquinaForm(request.POST, instance=maquina)
        if form.is_valid():
            form.save()
            messages.success(request, f'La máquina "{maquina.nombre}" ha sido actualizada exitosamente.')
            return redirect('ver_detalle_maquina', id=maquina.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MaquinaForm(instance=maquina)
    
    return render(request, 'editar_maquina.html', {
        'form': form,
        'maquina': maquina
    })

def comparar_maquina(request, id):
    """Vista para comparar una máquina"""
    maquina = get_object_or_404(Maquina, id=id)
    return render(request, 'comparar.html', {
        'maquina': maquina
    })

@require_http_methods(["GET"])
def api_maquinas_por_tipo(request, tipo):
    """API para obtener máquinas por tipo"""
    try:
        maquinas = Maquina.objects.filter(tipo=tipo).values(
            'id', 'nombre', 'numero_serie', 'tipo'
        )
        return JsonResponse(list(maquinas), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_http_methods(["GET"])
def api_maquina_detalle(request, id):
    """API para obtener detalles de una máquina"""
    try:
        maquina = get_object_or_404(Maquina, id=id)
        
        # Convertir el objeto máquina a diccionario
        maquina_dict = {
            'id': str(maquina.id),
            'nombre': maquina.nombre,
            'tipo': maquina.tipo,
            'numero_serie': maquina.numero_serie,
            'criticality_ranking': maquina.criticality_ranking,
            'availability': maquina.availability,
            'date_in_service': maquina.date_in_service.isoformat() if maquina.date_in_service else None,
            'purchase_price': float(maquina.purchase_price) if maquina.purchase_price else None,
            'installation_and_training_cost': float(maquina.installation_and_training_cost) if maquina.installation_and_training_cost else None,
            'setup_costs': float(maquina.setup_costs) if maquina.setup_costs else None,
            'current_resale_value': float(maquina.current_resale_value) if maquina.current_resale_value else None,
            'salvage_value': float(maquina.salvage_value) if maquina.salvage_value else None,
            'annual_maintenance_labor_parts': float(maquina.annual_maintenance_labor_parts) if maquina.annual_maintenance_labor_parts else None,
            'initial_monthly_maintenance_cost': float(maquina.initial_monthly_maintenance_cost) if maquina.initial_monthly_maintenance_cost else None,
            'maintenance_cost_gradient': maquina.maintenance_cost_gradient,
            'cost_of_downtime': float(maquina.cost_of_downtime) if maquina.cost_of_downtime else None,
            'operator_labor_cost': float(maquina.operator_labor_cost) if maquina.operator_labor_cost else None,
            'energy_consumption': maquina.energy_consumption,
            'energy_cost': float(maquina.energy_cost) if maquina.energy_cost else None,
            'consumable_replacement_cost_1': float(maquina.consumable_replacement_cost_1) if maquina.consumable_replacement_cost_1 else None,
            'consumable_lifespan_1': maquina.consumable_lifespan_1,
            'useful_life': maquina.useful_life,
            'end_of_useful_life_date': maquina.end_of_useful_life_date.isoformat() if maquina.end_of_useful_life_date else None,
            'monthly_depreciation': float(maquina.monthly_depreciation) if maquina.monthly_depreciation else None,
            'production_rate': maquina.production_rate,
            'production_rate_units': maquina.production_rate_units,
            'texas_workdays': maquina.texas_workdays,
            'monthly_operating_hours': maquina.monthly_operating_hours,
        }
        
        return JsonResponse(maquina_dict)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_http_methods(["GET"])
def api_comparar_maquinas(request, id1, id2):
    """API para comparar dos máquinas"""
    try:
        maquina1 = get_object_or_404(Maquina, id=id1)
        maquina2 = get_object_or_404(Maquina, id=id2)
        
        # Obtener todos los campos del modelo (excluyendo el ID)
        campos_modelo = [field.name for field in Maquina._meta.fields if field.name != 'id']
        
        # Obtener valores de ambas máquinas
        valores_maquina1 = {}
        valores_maquina2 = {}
        
        for campo in campos_modelo:
            valor1 = getattr(maquina1, campo)
            valor2 = getattr(maquina2, campo)
            
            # Convertir Decimal a float para JSON
            if hasattr(valor1, 'quantize'):  # Es un Decimal
                valor1 = float(valor1) if valor1 is not None else None
            if hasattr(valor2, 'quantize'):  # Es un Decimal
                valor2 = float(valor2) if valor2 is not None else None
            
            # Convertir fechas a string
            if hasattr(valor1, 'isoformat'):
                valor1 = valor1.isoformat() if valor1 else None
            if hasattr(valor2, 'isoformat'):
                valor2 = valor2.isoformat() if valor2 else None
            
            valores_maquina1[campo] = valor1
            valores_maquina2[campo] = valor2
        
        # Encontrar campos en común (ambos tienen valores no nulos)
        campos_comunes = []
        for campo in campos_modelo:
            if (valores_maquina1[campo] is not None and 
                valores_maquina2[campo] is not None and
                valores_maquina1[campo] != '' and 
                valores_maquina2[campo] != ''):
                campos_comunes.append(campo)
        
        # Calcular análisis de rentabilidad básico
        analisis_rentabilidad = calcular_analisis_rentabilidad(maquina1, maquina2)
        
        response_data = {
            'maquina1': {
                'id': str(maquina1.id),
                'nombre': maquina1.nombre,
                'tipo': maquina1.tipo,
                'valores': valores_maquina1
            },
            'maquina2': {
                'id': str(maquina2.id),
                'nombre': maquina2.nombre,
                'tipo': maquina2.tipo,
                'valores': valores_maquina2
            },
            'campos_comunes': campos_comunes,
            'total_campos': len(campos_modelo),
            'analisis_rentabilidad': analisis_rentabilidad
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def calcular_analisis_rentabilidad(maquina1, maquina2):
    """Calcular análisis básico de rentabilidad"""
    try:
        # Verificar que tengamos los datos mínimos necesarios
        if not all([
            maquina1.purchase_price, 
            maquina2.purchase_price,
            maquina1.annual_maintenance_labor_parts,
            maquina2.annual_maintenance_labor_parts
        ]):
            return None
        
        # Calcular diferencia en costo de adquisición
        diferencia_compra = float(maquina2.purchase_price - maquina1.purchase_price)
        
        # Calcular diferencia en mantenimiento anual
        diferencia_mantenimiento = float(maquina2.annual_maintenance_labor_parts - maquina1.annual_maintenance_labor_parts)
        
        # Calcular payback period simple (años)
        if diferencia_mantenimiento != 0:
            payback_years = abs(diferencia_compra / diferencia_mantenimiento)
        else:
            payback_years = float('inf')
        
        # Determinar recomendación
        if diferencia_compra < 0 and diferencia_mantenimiento < 0:
            recomendacion = "reemplazar"
            mensaje = f"Recomendado: La nueva máquina cuesta menos y tiene menor mantenimiento."
        elif diferencia_compra > 0 and diferencia_mantenimiento < 0:
            if payback_years <= 3:
                recomendacion = "reemplazar"
                mensaje = f"Recomendado: Se recupera la inversión en {payback_years:.1f} años."
            else:
                recomendacion = "evaluar"
                mensaje = f"Evaluar: Payback de {payback_years:.1f} años puede ser alto."
        else:
            recomendacion = "no_reemplazar"
            mensaje = "No recomendado: Mayor costo inicial y de mantenimiento."
        
        return {
            'recomendacion': recomendacion,
            'mensaje': mensaje,
            'diferencia_compra': diferencia_compra,
            'diferencia_mantenimiento_anual': diferencia_mantenimiento,
            'payback_years': payback_years if payback_years != float('inf') else None
        }
        
    except Exception as e:
        return None

def clientes_list(request):
    """Lista todos los clientes activos"""
    clientes = Cliente.objects.filter(activo=True)
    return render(request, 'listar_clientes.html', {'clientes': clientes})

def clientes_crear(request):
    """Crear nuevo cliente"""
    if request.method == 'POST':
        try:
            Cliente.objects.create(
                id=request.POST.get('id'),
                nombre=request.POST.get('nombre'),
                contacto=request.POST.get('contacto'),
                email=request.POST.get('email') or None,
                telefono=request.POST.get('telefono') or None,
                terminos_contractuales=int(request.POST.get('terminos_contractuales')),
                average_days_to_pay=int(request.POST.get('average_days_to_pay')),
                observaciones=request.POST.get('observaciones', '')
            )
            messages.success(request, 'Cliente creado exitosamente')
            return redirect('clientes_list')
        except Exception as e:
            messages.error(request, f'Error al crear cliente: {str(e)}')
    
    return render(request, 'crear_clientes.html')

def clientes_editar(request, pk):
    """Editar cliente existente"""
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if request.method == 'POST':
        try:
            cliente.nombre = request.POST.get('nombre')
            cliente.contacto = request.POST.get('contacto')
            cliente.email = request.POST.get('email') or None
            cliente.telefono = request.POST.get('telefono') or None
            cliente.terminos_contractuales = int(request.POST.get('terminos_contractuales'))
            cliente.average_days_to_pay = int(request.POST.get('average_days_to_pay'))
            cliente.observaciones = request.POST.get('observaciones', '')
            cliente.save()
            messages.success(request, 'Cliente actualizado exitosamente')
            return redirect('clientes_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar cliente: {str(e)}')
    
    return render(request, 'editar_clientes.html', {'cliente': cliente})

def clientes_eliminar(request, pk):
    """Eliminar cliente (desactivar)"""
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = False
    cliente.save()
    messages.success(request, 'Cliente desactivado exitosamente')
    return redirect('clientes_list')

# ==================== VISTAS DE PROVEEDORES ====================

def proveedores_list(request):
    """Lista todos los proveedores activos"""
    proveedores = Proveedor.objects.filter(activo=True)
    return render(request, 'listar_proveedores.html', {'proveedores': proveedores})

def proveedores_crear(request):
    """Crear nuevo proveedor"""
    if request.method == 'POST':
        try:
            Proveedor.objects.create(
                id=request.POST.get('id'),
                nombre=request.POST.get('nombre'),
                contacto=request.POST.get('contacto'),
                email=request.POST.get('email') or None,
                telefono=request.POST.get('telefono') or None,
                terminos_pago=int(request.POST.get('terminos_pago')),
                tipo_materia_prima=request.POST.get('tipo_materia_prima') or None,
                observaciones=request.POST.get('observaciones', '')
            )
            messages.success(request, 'Proveedor creado exitosamente')
            return redirect('proveedores_list')
        except Exception as e:
            messages.error(request, f'Error al crear proveedor: {str(e)}')
    
    return render(request, 'crear_proveedores.html')

def proveedores_editar(request, pk):
    """Editar proveedor existente"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        try:
            proveedor.nombre = request.POST.get('nombre')
            proveedor.contacto = request.POST.get('contacto')
            proveedor.email = request.POST.get('email') or None
            proveedor.telefono = request.POST.get('telefono') or None
            proveedor.terminos_pago = int(request.POST.get('terminos_pago'))
            proveedor.tipo_materia_prima = request.POST.get('tipo_materia_prima') or None
            proveedor.observaciones = request.POST.get('observaciones', '')
            proveedor.save()
            messages.success(request, 'Proveedor actualizado exitosamente')
            return redirect('proveedores_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar proveedor: {str(e)}')
    
    return render(request, 'editar_proveedores.html', {'proveedor': proveedor})

def proveedores_eliminar(request, pk):
    """Eliminar proveedor (desactivar)"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    proveedor.activo = False
    proveedor.save()
    messages.success(request, 'Proveedor desactivado exitosamente')
    return redirect('proveedores_list')

# ==================== VISTAS DE REGISTROS ====================

def registros_list(request):
    """
    Vista minimalista para listar registros. Los cálculos se delegan a JavaScript.
    """
    query = request.GET.get('q', '')
    
    # Query base optimizada que se ordena por defecto según el Meta del modelo
    registros_qs = Registro.objects.select_related('cliente')

    # Filtra solo si hay un término de búsqueda
    if query:
        registros_qs = registros_qs.filter(
            Q(id__icontains=query) |
            Q(cliente__nombre__icontains=query) |
            Q(estado_cobro__icontains=query)
        )

    context = {
        'registros': registros_qs,
        'query': query,
    }

    return render(request, 'listar_registros.html', context)

def registros_crear(request):
    """Vista para crear registro con gestión integral y cálculo automático de fechas."""
    
    # --- Lógica para la petición GET (Mostrar el formulario) ---
    if request.method != 'POST':
        form = RegistroForm()
        
        # 1. PREPARAR DATOS PARA EL FRONTEND (CAMBIO PRINCIPAL)
        # Se obtienen los términos de todos los clientes y proveedores para inyectarlos en el script.
        # Esto es más eficiente que hacer múltiples llamadas AJAX.
        
        clientes = Cliente.objects.filter(activo=True)
        clientes_data = {
            str(c.id): {
                'nombre': c.nombre,
                'terminos': c.terminos_contractuales
            } for c in clientes
        }

        proveedores = Proveedor.objects.filter(activo=True)
        proveedores_data = {
            str(p.id): {
                'nombre': p.nombre,
                'terminos': p.terminos_pago
            } for p in proveedores
        }

        context = {
            'form': form,
            'fecha_hoy': date.today().isoformat(),
            'metodos_pago': Registro.METODO_PAGO_CHOICES,
            # Convertir los diccionarios a JSON para que JS los pueda leer
            'clientes_data': json.dumps(clientes_data),
            'proveedores_data': json.dumps(proveedores_data),
        }
        # Asegúrate que el nombre del template sea el correcto
        return render(request, 'crear_registros.html', context)

    # --- Lógica para la petición POST (Guardar el formulario) ---
    form = RegistroForm(request.POST)
    
    obligaciones_data = request.POST.get('obligaciones_data', '[]')
    pagos_cliente_data = request.POST.get('pagos_cliente_data', '[]')
    pagos_proveedor_data = request.POST.get('pagos_proveedor_data', '[]')
    
    try:
        obligaciones = json.loads(obligaciones_data)
        pagos_cliente = json.loads(pagos_cliente_data)
        pagos_proveedor = json.loads(pagos_proveedor_data)
    except json.JSONDecodeError:
        messages.error(request, 'Error en el formato de datos JSON.')
        # Re-renderiza el contexto como lo haría la petición GET
        # (Esto es una mejora para que el formulario no pierda los datos de JS)
        # Puedes copiar la lógica del GET aquí si quieres que el formulario se repoble
        # con los datos de JS en caso de error. Por simplicidad, se omite por ahora.
        return redirect('registros_crear') # Redirigir es más simple

    if form.is_valid():
        try:
            with transaction.atomic():
                registro = form.save(commit=False)
                
                # 2. CALCULAR FECHAS DE VENCIMIENTO EN EL BACKEND (CAMBIO PRINCIPAL)
                # El frontend envía 'fecha_recepcion'. El backend calcula la 'fecha_vencimiento' final.
                # Esto garantiza que la lógica de negocio resida en el servidor.
                
                # Cache para proveedores para no consultar la DB en un bucle
                proveedores_cache = {str(p.id): p for p in Proveedor.objects.filter(
                    id__in=[obl.get('proveedor_id') for obl in obligaciones]
                )}

                obligaciones_procesadas = []
                for i, obligacion in enumerate(obligaciones):
                    proveedor_id = obligacion.get('proveedor_id')
                    fecha_recepcion_str = obligacion.get('fecha_recepcion')

                    if not proveedor_id or not fecha_recepcion_str:
                        raise ValidationError(f'La obligación #{i+1} debe tener proveedor y fecha de recepción.')
                    
                    proveedor = proveedores_cache.get(proveedor_id)
                    if not proveedor:
                        raise ValidationError(f'Proveedor con ID {proveedor_id} no encontrado.')

                    # Calcular la fecha de vencimiento real
                    fecha_recepcion = datetime.strptime(fecha_recepcion_str, '%Y-%m-%d').date()
                    fecha_vencimiento = fecha_recepcion + timedelta(days=proveedor.terminos_pago)
                    
                    obligacion_procesada = obligacion.copy()
                    obligacion_procesada['id'] = i + 1
                    obligacion_procesada['fecha_vencimiento'] = fecha_vencimiento.isoformat() # Guardar fecha calculada
                    obligaciones_procesadas.append(obligacion_procesada)

                # El resto de tu lógica de validación y procesamiento es correcta y se mantiene
                # ... (validaciones de pagos, asignación de IDs, etc.) ...
                total_pagos_cliente = Decimal('0.00')
                for pago in pagos_cliente:
                    total_pagos_cliente += Decimal(str(pago.get('monto', 0)))
                
                if total_pagos_cliente > registro.valor_cobrar_cliente:
                    raise ValidationError(f'Los pagos del cliente (${total_pagos_cliente:,.2f}) exceden el valor a cobrar (${registro.valor_cobrar_cliente:,.2f}).')

                # Procesar IDs
                pagos_cliente_procesados = [{'id': i + 1, **pago} for i, pago in enumerate(pagos_cliente) if pago.get('monto')]
                pagos_proveedor_procesados = [{'id': i + 1, **pago} for i, pago in enumerate(pagos_proveedor) if pago.get('monto')]
                
                # Asignar datos procesados
                registro.obligaciones_data = obligaciones_procesadas
                registro.pagos_cliente_data = pagos_cliente_procesados
                registro.pagos_proveedor_data = pagos_proveedor_procesados
                
                # El método save() del modelo se encargará de la fecha_limite_cobro del cliente
                registro.save() 
                if registro.cliente:
                    registro.cliente.actualizar_dias_promedio_pago()
                messages.success(request, f'Registro {registro.id} creado exitosamente.')
                return redirect('registros_list')
                
        except ValidationError as e:
            # Captura errores de validación de forma clara
            messages.error(request, e.message)
        except Exception as e:
            messages.error(request, f'Error inesperado al guardar: {str(e)}')
    else:
        # Si el formulario principal no es válido
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f'Error en el campo "{form.fields[field].label}": {error}')
    
    # Si algo falla, volver a renderizar el formulario
    # (Aquí también podrías repoblar el contexto para no perder datos)
    return redirect('registros_crear')

def registros_editar(request, id):
    """Vista para editar registro existente con gestión integral."""
    
    # Obtener el registro o devolver 404
    registro = get_object_or_404(Registro, id=id)  # Changed registro_id to id
    
    # --- Lógica para la petición GET (Mostrar el formulario con datos existentes) ---
    if request.method != 'POST':
        form = RegistroForm(instance=registro)
        
        # Preparar datos para el frontend
        clientes = Cliente.objects.filter(activo=True)
        clientes_data = {
            str(c.id): {
                'nombre': c.nombre,
                'terminos': c.terminos_contractuales
            } for c in clientes
        }

        proveedores = Proveedor.objects.filter(activo=True)
        proveedores_data = {
            str(p.id): {
                'nombre': p.nombre,
                'terminos': p.terminos_pago
            } for p in proveedores
        }

        # Preparar datos existentes del registro para el frontend
        registro_data = {
            'obligaciones': registro.obtener_obligaciones(),
            'pagos_cliente': registro.obtener_pagos_cliente(),
            'pagos_proveedor': registro.obtener_pagos_proveedor(),
        }

        context = {
            'form': form,
            'registro': registro,
            'fecha_hoy': date.today().isoformat(),
            'metodos_pago': Registro.METODO_PAGO_CHOICES,
            'clientes_data': json.dumps(clientes_data),
            'proveedores_data': json.dumps(proveedores_data),
            'registro_data': json.dumps(registro_data),
            'es_edicion': True,
        }
        return render(request, 'editar_registro.html', context)

    # --- Lógica para la petición POST (Actualizar el formulario) ---
    form = RegistroForm(request.POST, instance=registro)
    
    obligaciones_data = request.POST.get('obligaciones_data', '[]')
    pagos_cliente_data = request.POST.get('pagos_cliente_data', '[]')
    pagos_proveedor_data = request.POST.get('pagos_proveedor_data', '[]')
    
    try:
        obligaciones = json.loads(obligaciones_data)
        pagos_cliente = json.loads(pagos_cliente_data)
        pagos_proveedor = json.loads(pagos_proveedor_data)
    except json.JSONDecodeError:
        messages.error(request, 'Error en el formato de datos JSON.')
        return redirect('registros_editar', id=id)  # Changed registro_id to id

    if form.is_valid():
        try:
            with transaction.atomic():
                # Guardar los cambios básicos del formulario
                registro_actualizado = form.save(commit=False)
                
                # Procesar obligaciones
                proveedores_cache = {str(p.id): p for p in Proveedor.objects.filter(
                    id__in=[obl.get('proveedor_id') for obl in obligaciones if obl.get('proveedor_id')]
                )}

                obligaciones_procesadas = []
                for i, obligacion in enumerate(obligaciones):
                    proveedor_id = obligacion.get('proveedor_id')
                    fecha_recepcion_str = obligacion.get('fecha_recepcion')

                    if not proveedor_id or not fecha_recepcion_str:
                        raise ValidationError(f'La obligación #{i+1} debe tener proveedor y fecha de recepción.')
                    
                    proveedor = proveedores_cache.get(proveedor_id)
                    if not proveedor:
                        raise ValidationError(f'Proveedor con ID {proveedor_id} no encontrado.')

                    # Calcular la fecha de vencimiento
                    fecha_recepcion = datetime.strptime(fecha_recepcion_str, '%Y-%m-%d').date()
                    fecha_vencimiento = fecha_recepcion + timedelta(days=proveedor.terminos_pago)
                    
                    obligacion_procesada = obligacion.copy()
                    # Mantener ID existente o crear nuevo
                    if 'id' not in obligacion_procesada or not obligacion_procesada['id']:
                        obligacion_procesada['id'] = i + 1
                    obligacion_procesada['fecha_vencimiento'] = fecha_vencimiento.isoformat()
                    obligacion_procesada['proveedor_nombre'] = proveedor.nombre
                    obligaciones_procesadas.append(obligacion_procesada)

                # Validar pagos del cliente
                total_pagos_cliente = Decimal('0.00')
                for pago in pagos_cliente:
                    if pago.get('monto'):
                        total_pagos_cliente += Decimal(str(pago.get('monto', 0)))
                
                if total_pagos_cliente > registro_actualizado.valor_cobrar_cliente:
                    raise ValidationError(f'Los pagos del cliente (${total_pagos_cliente:,.2f}) exceden el valor a cobrar (${registro_actualizado.valor_cobrar_cliente:,.2f}).')

                # Procesar pagos manteniendo IDs existentes o creando nuevos
                pagos_cliente_procesados = []
                for i, pago in enumerate(pagos_cliente):
                    if pago.get('monto'):
                        pago_procesado = pago.copy()
                        if 'id' not in pago_procesado or not pago_procesado['id']:
                            # Generar nuevo ID basado en los existentes
                            max_id = max([p.get('id', 0) for p in pagos_cliente_procesados], default=0)
                            pago_procesado['id'] = max_id + 1
                        pagos_cliente_procesados.append(pago_procesado)
                
                pagos_proveedor_procesados = []
                for i, pago in enumerate(pagos_proveedor):
                    if pago.get('monto'):
                        pago_procesado = pago.copy()
                        if 'id' not in pago_procesado or not pago_procesado['id']:
                            max_id = max([p.get('id', 0) for p in pagos_proveedor_procesados], default=0)
                            pago_procesado['id'] = max_id + 1
                        pagos_proveedor_procesados.append(pago_procesado)
                
                # Actualizar datos del registro
                registro_actualizado.obligaciones_data = obligaciones_procesadas
                registro_actualizado.pagos_cliente_data = pagos_cliente_procesados
                registro_actualizado.pagos_proveedor_data = pagos_proveedor_procesados
                
                # Guardar registro actualizado
                registro_actualizado.save()
                
                # Actualizar estado de cobro automáticamente
                registro_actualizado.actualizar_estado_cobro()
                
                # Actualizar días promedio de pago del cliente
                if registro_actualizado.cliente:
                    registro_actualizado.cliente.actualizar_dias_promedio_pago()
                    
                messages.success(request, f'Registro {registro_actualizado.id} actualizado exitosamente.')
                return redirect('registros_list')
                
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error inesperado al actualizar: {str(e)}')
    else:
        # Si el formulario principal no es válido
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f'Error en el campo "{form.fields[field].label}": {error}')
    
    # Si algo falla, volver a mostrar el formulario
    return redirect('registros_editar', id=id)  # Changed registro_id to id


# Vista auxiliar para validar ID de registro
def validar_id_registro(request):
    """Vista AJAX para validar si un ID de registro ya existe"""
    if request.method == 'GET':
        id_registro = request.GET.get('id', '')
        existe = Registro.objects.filter(id=id_registro).exists()
        return JsonResponse({'existe': existe})

# Vista auxiliar para obtener términos de cliente
def obtener_terminos_cliente(request, cliente_id):
    """Vista AJAX para obtener términos de un cliente"""
    try:
        cliente = get_object_or_404(Cliente, id=cliente_id)
        return JsonResponse({
            'success': True,
            'terminos_contractuales': cliente.terminos_contractuales,
            'nombre': cliente.nombre
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Error al obtener datos del cliente'
        })

# Vista auxiliar para obtener términos de proveedor
def obtener_terminos_proveedor(request, proveedor_id):
    """Vista AJAX para obtener términos de un proveedor"""
    try:
        proveedor = get_object_or_404(Proveedor, id=proveedor_id)
        return JsonResponse({
            'success': True,
            'terminos_pago': getattr(proveedor, 'terminos_pago', 30),
            'nombre': proveedor.nombre
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'Error al obtener datos del proveedor'
        })

# ==================== FLUJO DE CAJA ====================

def flujo_caja_view(request, registro_id):
    """
    Vista principal para mostrar el flujo de caja detallado de un registro
    """
    registro = get_object_or_404(Registro, pk=registro_id)

    # Procesar obligaciones con información detallada
    obligaciones = []
    for obl in registro.obtener_obligaciones():
        pagos = registro.obtener_pagos_de_obligacion(obl.get('id'))
        pagado = sum([float(p.get('monto', 0)) for p in pagos])
        total = float(obl.get('valor_pagar', 0))
        saldo = total - pagado
        
        # Calcular días de vencimiento
        fecha_vencimiento = obl.get('fecha_vencimiento')
        dias_vencimiento = None
        if fecha_vencimiento:
            if isinstance(fecha_vencimiento, str):
                try:
                    fecha_vencimiento = datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date()
                except ValueError:
                    fecha_vencimiento = None
            
            if fecha_vencimiento:
                dias_vencimiento = (fecha_vencimiento - date.today()).days

        obligaciones.append({
            'id': obl.get('id'),
            'proveedor': obl.get('proveedor_nombre', 'N/A'),
            'valor_pagar': total,
            'pagado': pagado,
            'saldo': saldo,
            'fecha_vencimiento': fecha_vencimiento,
            'dias_vencimiento': dias_vencimiento,
            'descripcion': obl.get('descripcion', ''),
        })

    # Ordenar obligaciones por fecha de vencimiento
    obligaciones.sort(key=lambda x: x['fecha_vencimiento'] if x['fecha_vencimiento'] else date.max)

    context = {
        'registro': registro,
        'obligaciones': obligaciones,
    }
    return render(request, 'flujo_caja.html', context)

def calcular_flujo_caja(request):
    """
    Vista para calcular proyecciones de flujo de caja
    """
    registro_id = request.GET.get('registro_id')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    registro = get_object_or_404(Registro, id=registro_id)

    # Convertir fechas
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        fecha_inicio = registro.fecha_entrega_cliente

    try:
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        fecha_fin = date.today() + timezone.timedelta(days=30)  # 30 días hacia adelante por defecto

    # Proyección diaria de flujo
    proyecciones = registro.obtener_proyeccion_flujo(fecha_inicio, fecha_fin)

    # Agrupar ingresos y egresos por fecha
    flujo_por_fecha = {}
    for proyeccion in proyecciones:
        fecha = proyeccion['fecha']
        monto = Decimal(proyeccion['monto'])
        tipo = proyeccion['tipo']
        concepto = proyeccion['concepto']

        if fecha not in flujo_por_fecha:
            flujo_por_fecha[fecha] = {
                'fecha': fecha.isoformat(),
                'ingresos_esperados': 0,
                'egresos_esperados': 0,
                'detalles': {
                    'ingresos': [],
                    'egresos': []
                }
            }

        if tipo == 'ingreso':
            flujo_por_fecha[fecha]['ingresos_esperados'] += float(monto)
            flujo_por_fecha[fecha]['detalles']['ingresos'].append({
                'cliente': registro.cliente.nombre,
                'concepto': concepto,
                'monto': float(monto)
            })
        else:
            flujo_por_fecha[fecha]['egresos_esperados'] += float(monto)
            flujo_por_fecha[fecha]['detalles']['egresos'].append({
                'proveedor': concepto.replace("Pago a ", ""),
                'concepto': concepto,
                'monto': float(monto)
            })

    # Ordenar por fecha y calcular flujo acumulado
    fechas_ordenadas = sorted(flujo_por_fecha.keys())
    flujo_ordenado = []
    flujo_acumulado = 0
    dias_con_flujo_negativo = 0

    for fecha in fechas_ordenadas:
        dia = flujo_por_fecha[fecha]
        flujo_neto = dia['ingresos_esperados'] - dia['egresos_esperados']
        flujo_acumulado += flujo_neto
        
        if flujo_acumulado < 0:
            dias_con_flujo_negativo += 1

        dia['flujo_neto'] = flujo_neto
        dia['flujo_acumulado'] = flujo_acumulado
        flujo_ordenado.append(dia)

    # Calcular resumen estadístico
    resumen = {
        'total_ingresos': sum(d['ingresos_esperados'] for d in flujo_ordenado),
        'total_egresos': sum(d['egresos_esperados'] for d in flujo_ordenado),
        'flujo_neto_total': sum(d['flujo_neto'] for d in flujo_ordenado),
        'flujo_acumulado_min': min((d['flujo_acumulado'] for d in flujo_ordenado), default=0),
        'flujo_acumulado_max': max((d['flujo_acumulado'] for d in flujo_ordenado), default=0),
        'dias_con_flujo_negativo': dias_con_flujo_negativo,
        'periodo_dias': (fecha_fin - fecha_inicio).days + 1,
        'promedio_diario': sum(d['flujo_neto'] for d in flujo_ordenado) / len(flujo_ordenado) if flujo_ordenado else 0
    }

    return JsonResponse({
        'success': True,
        'data': flujo_ordenado,
        'resumen': resumen,
        'registro': {
            'id': registro.id,
            'cliente': registro.cliente.nombre,
            'valor_cobrar': float(registro.valor_cobrar_cliente),
            'saldo_pendiente': float(registro.calcular_saldo_pendiente_cliente()),
        }
    })

def obtener_datos_dashboard(request, registro_id):
    """
    Vista para obtener datos del dashboard en tiempo real (AJAX)
    """
    registro = get_object_or_404(Registro, pk=registro_id)
    
    # Calcular métricas principales
    valor_cobrar = float(registro.valor_cobrar_cliente)
    saldo_pendiente = float(registro.calcular_saldo_pendiente_cliente())
    total_obligaciones = float(registro.calcular_total_obligaciones())
    margen_bruto = float(registro.margen_bruto) if hasattr(registro, 'margen_bruto') else 0
    
    # Calcular porcentajes
    porcentaje_cobrado = ((valor_cobrar - saldo_pendiente) / valor_cobrar * 100) if valor_cobrar > 0 else 0
    
    # Calcular porcentaje pagado a proveedores
    pagado_proveedores = 0
    for obl in registro.obtener_obligaciones():
        pagos = registro.obtener_pagos_de_obligacion(obl.get('id'))
        pagado_proveedores += sum([float(p.get('monto', 0)) for p in pagos])
    
    porcentaje_pagado_proveedores = (pagado_proveedores / total_obligaciones * 100) if total_obligaciones > 0 else 0
    
    # Análisis de riesgo
    riesgo = registro.analizar_riesgo_cobro() if hasattr(registro, 'analizar_riesgo_cobro') else {
        'nivel': 'bajo',
        'mensaje': 'Análisis de riesgo no disponible'
    }
    
    return JsonResponse({
        'success': True,
        'metricas': {
            'valor_cobrar': valor_cobrar,
            'saldo_pendiente': saldo_pendiente,
            'total_obligaciones': total_obligaciones,
            'margen_bruto': margen_bruto,
            'porcentaje_cobrado': round(porcentaje_cobrado, 1),
            'porcentaje_pagado_proveedores': round(porcentaje_pagado_proveedores, 1),
        },
        'riesgo': riesgo,
        'timeline': {
            'pagos_cliente': len(registro.obtener_pagos_cliente()),
            'pagos_proveedor': len(registro.obtener_pagos_proveedor()),
            'obligaciones_total': len(registro.obtener_obligaciones()),
        }
    })

def exportar_reporte_flujo(request, registro_id):
    """
    Vista para exportar el reporte de flujo de caja
    """
    registro = get_object_or_404(Registro, pk=registro_id)
    
    # Aquí implementarías la lógica para generar PDF, Excel, etc.
    # Por ahora retornamos un JSON con los datos
    
    obligaciones = []
    for obl in registro.obtener_obligaciones():
        pagos = registro.obtener_pagos_de_obligacion(obl.get('id'))
        pagado = sum([float(p.get('monto', 0)) for p in pagos])
        total = float(obl.get('valor_pagar', 0))
        
        obligaciones.append({
            'proveedor': obl.get('proveedor_nombre', 'N/A'),
            'valor_total': total,
            'pagado': pagado,
            'saldo': total - pagado,
            'fecha_vencimiento': str(obl.get('fecha_vencimiento', '')),
            'descripcion': obl.get('descripcion', ''),
        })
    
    data = {
        'registro': {
            'id': registro.id,
            'cliente': registro.cliente.nombre,
            'fecha_entrega': str(registro.fecha_entrega_cliente),
            'valor_cobrar': float(registro.valor_cobrar_cliente),
            'saldo_pendiente': float(registro.calcular_saldo_pendiente_cliente()),
        },
        'obligaciones': obligaciones,
        'pagos_cliente': [
            {
                'fecha': str(p.get('fecha_pago', '')),
                'monto': float(p.get('monto', 0)),
                'metodo': p.get('metodo_pago', ''),
                'referencia': p.get('referencia', ''),
            }
            for p in registro.obtener_pagos_cliente()
        ],
        'pagos_proveedor': [
            {
                'fecha': str(p.get('fecha_pago', '')),
                'monto': float(p.get('monto', 0)),
                'metodo': p.get('metodo_pago', ''),
                'obligacion_id': p.get('obligacion_id', ''),
            }
            for p in registro.obtener_pagos_proveedor()
        ],
    }
    
    return JsonResponse({
        'success': True,
        'data': data,
        'message': 'Reporte generado exitosamente'
    })

# ================= IMPORTAR REGISTROS ==================

def cargar_excel_completo(request):
    if request.method == "POST" and request.FILES.get("archivo_excel"):
        archivo = request.FILES["archivo_excel"]
        try:
            wb = openpyxl.load_workbook(archivo)

            # === 1. CLIENTES ===
            if "Clientes" in wb.sheetnames:
                hoja_clientes = wb["Clientes"]
                for fila in list(hoja_clientes.iter_rows(min_row=2, values_only=True)):
                    (id_cliente, nombre, city, email, telefono,
                     terminos, avg_dias, activo, observaciones) = fila
                    if not Cliente.objects.filter(id=id_cliente).exists():
                        Cliente.objects.create(
                            id=id_cliente,
                            nombre=nombre,
                            city=city,
                            email=email,
                            telefono=telefono,
                            terminos_contractuales=int(terminos),
                            average_days_to_pay=int(avg_dias),
                            activo=bool(activo),
                            observaciones=observaciones or ""
                        )

            # === 2. PROVEEDORES ===
            if "Proveedores" in wb.sheetnames:
                hoja_prov = wb["Proveedores"]
                for fila in list(hoja_prov.iter_rows(min_row=2, values_only=True)):
                    (id_prov, nombre, contacto, email, telefono,
                     terminos, tipo_mp, activo, observaciones) = fila
                    if not Proveedor.objects.filter(id=id_prov).exists():
                        Proveedor.objects.create(
                            id=id_prov,
                            nombre=nombre,
                            contacto=contacto,
                            email=email,
                            telefono=telefono,
                            terminos_pago=int(terminos),
                            tipo_materia_prima=tipo_mp,
                            activo=bool(activo),
                            observaciones=observaciones or ""
                        )

            # === 3. REGISTROS ===
            hoja_registros = wb["Registros"]
            for fila in list(hoja_registros.iter_rows(min_row=2, values_only=True)):
                id_registro, cliente_id, fecha_entrega, valor_cobrar, estado_cobro, observaciones = fila
                try:
                    cliente = Cliente.objects.get(id=cliente_id)
                    if not Registro.objects.filter(id=id_registro).exists():
                        Registro.objects.create(
                            id=id_registro,
                            cliente=cliente,
                            fecha_entrega_cliente=fecha_entrega if isinstance(fecha_entrega, datetime) else datetime.strptime(fecha_entrega, "%Y-%m-%d").date(),
                            valor_cobrar_cliente=Decimal(valor_cobrar),
                            estado_cobro=estado_cobro,
                            observaciones=observaciones or ""
                        )
                except Exception as e:
                    messages.error(request, f"Registro {id_registro}: {str(e)}")

            # === 4. OBLIGACIONES ===
            hoja_obl = wb["Obligaciones"]
            for fila in list(hoja_obl.iter_rows(min_row=2, values_only=True)):
                registro_id, proveedor_id, proveedor_nombre, valor_pagar, fecha_vencimiento, descripcion, referencia = fila
                try:
                    registro = Registro.objects.get(id=registro_id)
                    registro.agregar_obligacion(
                        proveedor_nombre=proveedor_nombre,
                        valor_pagar=Decimal(valor_pagar),
                        fecha_vencimiento=fecha_vencimiento,
                        proveedor_id=proveedor_id,
                        descripcion=descripcion or "",
                        referencia=referencia or ""
                    )
                except Exception as e:
                    messages.error(request, f"Obligación para registro {registro_id}: {str(e)}")

            # === 5. PAGOS CLIENTE ===
            hoja_pc = wb["Pagos_Cliente"]
            for fila in list(hoja_pc.iter_rows(min_row=2, values_only=True)):
                registro_id, monto, fecha_pago, metodo_pago, referencia, observaciones = fila
                try:
                    registro = Registro.objects.get(id=registro_id)
                    registro.agregar_pago_cliente(
                        monto=Decimal(monto),
                        fecha_pago=fecha_pago,
                        metodo_pago=metodo_pago or "transferencia",
                        referencia=referencia or "",
                        observaciones=observaciones or ""
                    )
                except Exception as e:
                    messages.error(request, f"Pago cliente {registro_id}: {str(e)}")

            # === 6. PAGOS PROVEEDOR ===
            hoja_pp = wb["Pagos_Proveedor"]
            for fila in list(hoja_pp.iter_rows(min_row=2, values_only=True)):
                registro_id, obligacion_id, monto, fecha_pago, metodo_pago, referencia, observaciones = fila
                try:
                    registro = Registro.objects.get(id=registro_id)
                    registro.agregar_pago_proveedor(
                        obligacion_id=int(obligacion_id),
                        monto=Decimal(monto),
                        fecha_pago=fecha_pago,
                        metodo_pago=metodo_pago or "transferencia",
                        referencia=referencia or "",
                        observaciones=observaciones or ""
                    )
                except Exception as e:
                    messages.error(request, f"Pago proveedor {registro_id}: {str(e)}")

            messages.success(request, "Archivo importado correctamente.")

        except Exception as e:
            messages.error(request, f"Error general: {str(e)}")

        return redirect("cargar_excel_completo")

    return render(request, "cargar_excel_completo.html")












    